import os
import fitz
import re
import pymysql
from sentence_transformers import SentenceTransformer, CrossEncoder, util
from sympy import sympify, simplify
from textblob import TextBlob
from openpyxl import load_workbook

# ✅ Set HuggingFace cache paths
os.environ['HF_HOME'] = 'C:/ai_cache/huggingface'
os.environ['SENTENCE_TRANSFORMERS_HOME'] = 'C:/ai_cache/sentence_transformers'

# ✅ Load models
model = SentenceTransformer('paraphrase-MiniLM-L6-v2')
bi_encoder = SentenceTransformer('all-MiniLM-L6-v2')
cross_encoder = CrossEncoder('cross-encoder/stsb-roberta-large')

def db_connection():
    return pymysql.connect(host='localhost', user='root', password='', database='oop')

def evaluate_assignment(student_id, pdf_file):
    print(pdf_file)
    full_text = ""
    with fitz.open(pdf_file) as pdf:
        for page in pdf:
            full_text += page.get_text()

    answers = re.split(r"\?", full_text)
    print("Similarity scores:")
    # Fetch student questions
    student_db = pymysql.connect(host='localhost', user='root', password='', database='students')
    student_cursor = student_db.cursor()
    student_cursor.execute(f"SELECT q1, q2, q3, q4, q5 FROM details WHERE id='{student_id}'")
    student_questions = student_cursor.fetchone()
    student_db.close()

    if not student_questions:
        return {'error': 'Student not found in database'}

    # Fetch correct answers for those questions
    connection = db_connection()
    cursor = connection.cursor()
    question_answers = {}

    for q in student_questions:
        cursor.execute(f"SELECT question, answer FROM assignment_1 WHERE id={q}")
        row = cursor.fetchone()
        if row:
            question, answer = row
            question_answers[q] = {'question': question, 'answer': answer}

    connection.close()

    total_score = 0
    results = []

    # Evaluation criteria
    min_length_ratio = 0.6
    partial_credit_thresholds = [0.9, 0.8, 0.7]
    partial_scores = [10, 9, 8]
    j = 1
    x_scores=[]
    for i, (q_num, data) in enumerate(question_answers.items(), start=1):
        std_answer = answers[i].strip()
        correct_answer = data['answer']
        if(i < 5):
            next_q_num = list(question_answers.keys())[i]  # i is 1-based
            next_question_text = question_answers[next_q_num]['question']
            correct_answer += " " + next_question_text  # Append next question text to current answer
        bi_sim = util.cos_sim(
            bi_encoder.encode(correct_answer),
            bi_encoder.encode(std_answer)
        ).item()
        cross_sim = cross_encoder.predict([(correct_answer, std_answer)]) 
        similarity = 0.7 * cross_sim + 0.3 * bi_sim
        print(similarity)
        # If similarity is very low (< 0.3), treat it as irrelevant answer
        if similarity < 0.5:
            final_score = 2  # Completely wrong or off-topic
        else:
            # Length check
            length_score = 10 if len(std_answer) >= len(correct_answer) * min_length_ratio else 7

            # Key point coverage
            key_points = correct_answer.split(",")
            covered_points = sum(1 for point in key_points if point.strip().lower() in std_answer.lower())
            coverage_score = ((covered_points / len(key_points)) * 10) if key_points else 10

            # Formula correctness
            try:
                correct_formula = sympify(correct_answer)
                student_formula = sympify(std_answer)
                formula_score = 10 if simplify(correct_formula - student_formula) == 0 else 5
            except:
                formula_score = 0  # Set to 0 if not formula-based

            # Writing style bonus
            polarity = TextBlob(std_answer).sentiment.polarity
            style_bonus = 1 if polarity > 0 else 0

            # Base semantic score
            score = round(float(similarity) * 10)

            # Override score if in partial thresholds
            for threshold, partial_score in zip(partial_credit_thresholds, partial_scores):
                if similarity >= threshold:
                    score = partial_score
                    break

            # Final weighted score
            final_score = round(
                (0.75 * score) +
                (0.1 * length_score) +
                (0.1 * coverage_score) +
                (0.05 * formula_score) +
                style_bonus
            )

            final_score = max(1, min(10, final_score))  # Clamp between 1–10


        x_scores.append(final_score)
        final_score
        results.append({
            'question': j,
            'score': final_score
        })
        j += 1
        if final_score>5:
            total_score += 1

    print("Total Score:", total_score)

    # ✅ Write result to marks.xlsx
    try:
        excel_path = 'marks.xlsx'
        wb = load_workbook(excel_path)
        ws = wb.active

        found = False
        for row in ws.iter_rows(min_row=2, max_col=6):
            if row[0].value == student_id:
                row[1].value = total_score
                for i in range(2,6):
                    row[i].value=x_scores[i-2]
                found = True
                break

        if not found:
            ws.append([student_id, total_score]+x_scores)

        wb.save(excel_path)
        print(f"✅ Marks saved/updated for {student_id} in Excel.")
    except Exception as e:
        print(f"❌ Excel error: {e}")

    return results
